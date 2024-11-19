from openpyxl import Workbook 
from openpyxl.styles import PatternFill
 # The provided dictionary 
data = { "fc2709": ["H5", "I5", "P5", "Q5", "G6", "J6", "O6", "R6", "F7", "K7", "N7", "S7", "E8", "L8", "M8", "T8", "D9", "U9", "C10", "V10", "B11", "W11", "A12", "X12", "A13", "X13", "A14", "X14", "A15", "X15", "B16", "W16", "C17", "V17", "D18", "U18", "E19", "T19", "F20", "S20", "G21", "R21", "H22", "Q22", "I23", "P23", "J24", "O24", "K25", "N25", "L26", "M26"], "03ff00": ["H6", "I6", "P6", "Q6", "G7", "H7", "I7", "J7", "O7", "P7", "Q7", "R7", "F8", "G8", "J8", "K8", "N8", "O8", "R8", "S8", "E9", "F9", "K9", "L9", "M9", "N9", "S9", "T9", "D10", "E10", "T10", "U10", "C11", "D11", "U11", "V11", "B12", "C12", "V12", "W12", "B13", "C13", "V13", "W13", "B14", "C14", "V14", "W14", "B15", "C15", "V15", "W15", "C16", "D16", "U16", "V16", "D17", "E17", "T17", "U17", "E18", "F18", "S18", "T18", "F19", "G19", "R19", "S19", "G20", "H20", "Q20", "R20", "H21", "I21", "P21", "Q21", "I22", "J22", "O22", "P22", "J23", "K23", "N23", "O23", "K24", "L24", "M24", "N24", "L25", "M25"], "1b00ff": ["H8", "I8", "P8", "Q8", "G9", "H9", "I9", "J9", "O9", "P9", "Q9", "R9", "F10", "G10", "H10", "I10", "J10", "K10", "L10", "M10", "N10", "O10", "P10", "Q10", "R10", "S10", "E11", "F11", "G11", "H11", "K11", "L11", "M11", "N11", "O11", "R11", "S11", "T11", "D12", "E12", "F12", "S12", "T12", "U12", "D13", "E13", "F13", "S13", "T13", "U13", "D14", "E14", "F14", "S14", "T14", "U14", "D15", "E15", "F15", "S15", "T15", "U15", "E16", "F16", "S16", "T16", "F17", "G17", "R17", "S17", "G18", "H18", "Q18", "R18", "H19", "I19", "P19", "Q19", "I20", "J20", "O20", "P20", "J21", "K21", "N21", "O21", "K22", "L22", "M22", "N22", "L23", "M23"], "6400c4": ["I11", "J11", "P11", "Q11", "G12", "H12", "I12", "J12", "K12", "N12", "O12", "P12", "Q12", "R12", "G13", "I13", "J13", "K13", "L13", "M13", "N13", "O13", "Q13", "R13", "G14", "K14", "L14", "M14", "N14", "Q14", "R14", "G15", "Q15", "R15", "G16", "Q16", "R16", "H17", "P17", "Q17", "I18", "O18", "P18", "J19", "N19", "O19", "K20", "L20", "M20", "N20", "L21", "M21"], "C8C4C2": ["H13", "P13", "H14", "I14", "J14", "O14", "P14", "H15", "I15", "J15", "K15", "N15", "O15", "P15", "H16", "I16", "J16", "K16", "N16", "O16", "P16", "I17", "J17", "K17", "L17", "M17", "N17", "O17", "J18", "K18", "L18", "M18", "N18", "K19", "L19", "M19"] } 

wb = Workbook()
sheet = wb.active 

for i in range(1, 30):

    sheet.row_dimensions[i].height = 12 
    sheet.column_dimensions[sheet.cell(row=1, column=i).column_letter].width = 2

for color, cells in data.items():
    fill = PatternFill(start_color=color, end_color=color, fill_type="solid") 
    for cell in cells: 
        sheet[cell].fill = fill

wb.save("pixel_art.xlsx")